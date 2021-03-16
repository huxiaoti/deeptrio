# -*- coding: utf-8 -*-
"""
Created on Tue Mar  2 15:02:53 2021

@author: zju
"""

from openpyxl import load_workbook
wb = load_workbook(filename="pone.0125811.s001.xlsx")

sheets = wb.sheetnames
ws = wb['Positive PPIs Dataset']
ws2 = wb['Negative PPIs Dataset']

rows_1 = ws.rows
rows_2 = ws2.rows

you_name_1 = []
you_name_2 = []
you_list = []
you_seq = {}

for row in rows_1:
#    if row[1].value == '':
#        you_name_1.append(row[0].value)
#    else:
#        you_name_1.append(row[1].value)
    you_name_1.append(row[3].value)
#    if row[7].value == '':
#        you_name_2.append(row[6].value)
#    else:
#        you_name_2.append(row[7].value)
    you_name_2.append(row[9].value)
    you_seq[row[3].value] = row[4].value
    you_seq[row[9].value] = row[10].value

for k in you_name_1:
    if k not in you_list:
        you_list.append(k)
        
for k in you_name_2:
    if k not in you_list:
        you_list.append(k)

you_nega_1 = []
you_nega_2 = []

headlist = 0   
for row in rows_2:
    if headlist != 0:
        you_nega_1.append(row[0].value)
        you_nega_2.append(row[1].value)
    headlist = 1

wb2 = load_workbook(filename="pone.0125811.s002.xlsx")
ws3 = wb2['Protein List']
rows_3 = ws3.rows
for row in rows_3:
    if row[4].value != 'NaN':
        you_seq[row[4].value] = row[5].value
        
with open('loss_data.txt') as r:
    line = r.readline()
    while line != '':
        name = line.strip()
        seq = r.readline().strip()
        you_seq[int(name)] = seq
        line = r.readline()
    
    
you_number = []
with open('you_11188.txt', 'w') as w:
    for k in range(len(you_name_1)):
        if (len(you_seq[you_name_1[k]]) <= 1500) and (len(you_seq[you_name_2[k]]) <= 1500):
            w.write(str(you_name_1[k]) + '\t' + str(you_name_2[k]) + '\t1\n')
    for k in range(len(you_nega_1)):
        if (len(you_seq[you_nega_1[k]]) <= 1500) and (len(you_seq[you_nega_2[k]]) <= 1500):
            w.write(str(you_nega_1[k]) + '\t' + str(you_nega_2[k]) + '\t0\n')
                
you_number = list(set(you_number))

with open('you_database.txt', 'w') as w:
    for k in list(you_seq.keys()):
        w.write(str(k) + '\t' + you_seq[k] + '\n')